"""
Reporting and export service for CRM (Phase 1 - Post-deployment).
Generates PDF and Excel reports for clients, interactions, and performance metrics.
"""

import io
from datetime import datetime, timedelta
from django.db.models import Count, Q, Avg, Sum
from django.contrib.auth import get_user_model
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils import get_column_letter

from apps.crm.models import ClientProfile, ClientInteraction, PropertyInterest, Lead, ClientNote
from apps.crm.services.scope import client_profiles_for_user, reservations_for_user
from apps.auth.models import Agency

User = get_user_model()


def _end_of_day(dt):
    if dt and getattr(dt, 'hour', 0) == 0 and getattr(dt, 'minute', 0) == 0:
        return dt.replace(hour=23, minute=59, second=59, microsecond=999999)
    return dt


class ReportingService:
    """
    Service for generating reports and exports in PDF and Excel formats.
    """
    
    @staticmethod
    def generate_client_report_pdf(client_id, include_interactions=True, include_notes=True):
        """
        Generate a comprehensive PDF report for a specific client.
        
        Args:
            client_id: UUID of the client
            include_interactions: Whether to include interaction history
            include_notes: Whether to include client notes
            
        Returns:
            BytesIO object containing the PDF
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30,
                              topMargin=30, bottomMargin=30)
        
        # Container for the 'Flowable' objects
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a365d'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2c5282'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Get client data
        try:
            client_profile = ClientProfile.objects.select_related('user').get(id=client_id)
        except ClientProfile.DoesNotExist:
            # Return empty buffer if client doesn't exist
            return buffer
        
        user = client_profile.user
        
        # Title
        title = Paragraph(f"Rapport Client: {user.get_full_name()}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Client Information Section
        elements.append(Paragraph("Informations Générales", heading_style))
        
        client_data = [
            ['Email:', user.email],
            ['Téléphone:', getattr(user, 'phone', 'N/A')],
            ['Statut:', client_profile.get_status_display()],
            ['Niveau de priorité:', client_profile.get_priority_level_display()],
            ['Budget:', f"{client_profile.min_budget or 0} - {client_profile.max_budget or 'N/A'} FCFA"],
            ['Score de conversion:', f"{client_profile.conversion_score:.1f}%"],
            ['Date d\'inscription:', client_profile.created_at.strftime('%d/%m/%Y')],
        ]
        
        if client_profile.tags:
            client_data.append(['Tags:', ', '.join(client_profile.tags)])
        
        client_table = Table(client_data, colWidths=[2*inch, 4*inch])
        client_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e2e8f0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1a202c')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        elements.append(client_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Preferences Section
        elements.append(Paragraph("Préférences de Propriété", heading_style))
        
        pref_data = []
        if client_profile.preferred_property_types:
            pref_data.append(['Types préférés:', ', '.join(client_profile.preferred_property_types)])
        if client_profile.preferred_locations:
            pref_data.append(['Localisations:', ', '.join(client_profile.preferred_locations)])
        if client_profile.min_bedrooms or client_profile.max_bedrooms:
            bedrooms = f"{client_profile.min_bedrooms or 'N/A'} - {client_profile.max_bedrooms or 'N/A'}"
            pref_data.append(['Chambres:', bedrooms])
        if client_profile.min_area or client_profile.max_area:
            area = f"{client_profile.min_area or 'N/A'} - {client_profile.max_area or 'N/A'} m²"
            pref_data.append(['Surface:', area])
        
        if pref_data:
            pref_table = Table(pref_data, colWidths=[2*inch, 4*inch])
            pref_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e2e8f0')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1a202c')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(pref_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Property Interests
        interests = PropertyInterest.objects.filter(client=user).select_related('property')[:10]
        if interests.exists():
            elements.append(Paragraph("Historique des Intérêts (10 derniers)", heading_style))
            
            interest_data = [['Propriété', 'Type', 'Niveau d\'intérêt', 'Date']]
            for interest in interests:
                prop = interest.property
                interest_data.append([
                    prop.title[:40],
                    interest.get_interaction_type_display(),
                    interest.get_interest_level_display(),
                    interest.interaction_date.strftime('%d/%m/%Y')
                ])
            
            interest_table = Table(interest_data, colWidths=[2.5*inch, 1.2*inch, 1.3*inch, 1*inch])
            interest_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5282')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(interest_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Interactions
        if include_interactions:
            interactions = ClientInteraction.objects.filter(
                client=user
            ).select_related('agent').order_by('-scheduled_date')[:15]
            
            if interactions.exists():
                elements.append(PageBreak())
                elements.append(Paragraph("Historique des Interactions (15 dernières)", heading_style))
                
                interaction_data = [['Date', 'Type', 'Agent', 'Statut', 'Résultat']]
                for interaction in interactions:
                    interaction_data.append([
                        interaction.scheduled_date.strftime('%d/%m/%Y') if interaction.scheduled_date else 'N/A',
                        interaction.get_interaction_type_display(),
                        interaction.agent.get_full_name() if interaction.agent else 'N/A',
                        interaction.get_status_display(),
                        interaction.get_outcome_display() if interaction.outcome else 'N/A'
                    ])
                
                interaction_table = Table(interaction_data, colWidths=[1*inch, 1.2*inch, 1.5*inch, 1*inch, 1.3*inch])
                interaction_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5282')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                ]))
                elements.append(interaction_table)
        
        # Client Notes
        if include_notes:
            notes = ClientNote.objects.filter(
                client_profile=client_profile
            ).select_related('author').order_by('-created_at')[:10]
            
            if notes.exists():
                elements.append(PageBreak())
                elements.append(Paragraph("Notes Internes (10 dernières)", heading_style))
                
                for note in notes:
                    note_title = f"{note.title or 'Note sans titre'} - {note.get_note_type_display()}"
                    if note.is_important:
                        note_title += " [IMPORTANT]"
                    
                    elements.append(Paragraph(f"<b>{note_title}</b>", styles['Heading3']))
                    elements.append(Paragraph(
                        f"<i>Par {note.author.get_full_name()} le {note.created_at.strftime('%d/%m/%Y à %H:%M')}</i>",
                        styles['Normal']
                    ))
                    elements.append(Paragraph(note.content, styles['Normal']))
                    elements.append(Spacer(1, 0.2*inch))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generate_agent_performance_excel(agent_id, start_date=None, end_date=None):
        """
        Generate an Excel report with agent performance metrics.
        
        Args:
            agent_id: UUID of the agent
            start_date: Start date for the report period
            end_date: End date for the report period
            
        Returns:
            BytesIO object containing the Excel file
        """
        if not start_date:
            start_date = datetime.now() - timedelta(days=30)
        if not end_date:
            end_date = datetime.now()
        end_date = _end_of_day(end_date)
        
        # Get agent
        try:
            agent = User.objects.get(id=agent_id)
        except User.DoesNotExist:
            return io.BytesIO()
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Style definitions
        header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=14, color="1A365D")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Résumé"
        
        # Title
        ws_summary['A1'] = f"Rapport de Performance - {agent.get_full_name()}"
        ws_summary['A1'].font = title_font
        ws_summary['A2'] = f"Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        ws_summary.merge_cells('A1:D1')
        ws_summary.merge_cells('A2:D2')
        
        interactions = ClientInteraction.objects.filter(
            agent=agent,
            created_at__gte=start_date,
            created_at__lte=end_date,
        )
        
        total_interactions = interactions.count()
        completed_interactions = interactions.filter(status='completed').count()
        
        clients_managed = client_profiles_for_user(agent).count()
        new_clients = client_profiles_for_user(agent).filter(
            created_at__gte=start_date,
            created_at__lte=end_date,
        ).count()
        reservations_completed = reservations_for_user(agent).filter(
            status='completed',
            completed_at__gte=start_date,
            completed_at__lte=end_date,
        ).count()
        
        # Leads
        leads_assigned = Lead.objects.filter(
            assigned_agent=agent,
            created_at__gte=start_date,
            created_at__lte=end_date
        ).count()
        
        leads_converted = Lead.objects.filter(
            assigned_agent=agent,
            converted_to_client=True,
            conversion_date__gte=start_date,
            conversion_date__lte=end_date
        ).count()
        
        property_interests = PropertyInterest.objects.filter(
            client__interactions__agent=agent,
            interaction_date__gte=start_date,
            interaction_date__lte=end_date,
        ).distinct().count()
        
        try:
            from apps.commissions.models import Commission
            commission_total = Commission.objects.filter(
                agent=agent,
                transaction_date__gte=start_date,
                transaction_date__lte=end_date,
            ).aggregate(total=Sum('commission_amount'))['total'] or 0
        except Exception:
            commission_total = 0
        
        # Metrics table
        ws_summary['A4'] = "Métrique"
        ws_summary['B4'] = "Valeur"
        ws_summary['A4'].font = header_font
        ws_summary['B4'].font = header_font
        ws_summary['A4'].fill = header_fill
        ws_summary['B4'].fill = header_fill
        
        metrics = [
            ("Interactions totales", total_interactions),
            ("Interactions complétées", completed_interactions),
            ("Taux de complétion", f"{(completed_interactions/total_interactions*100):.1f}%" if total_interactions > 0 else "0%"),
            ("Clients (périmètre agence)", clients_managed),
            ("Nouveaux clients (période)", new_clients),
            ("Réservations terminées", reservations_completed),
            ("Commissions (FCFA)", float(commission_total)),
            ("Leads assignés", leads_assigned),
            ("Leads convertis", leads_converted),
            ("Taux de conversion leads", f"{(leads_converted/leads_assigned*100):.1f}%" if leads_assigned > 0 else "0%"),
            ("Intérêts propriétés", property_interests),
        ]
        
        row = 5
        for metric, value in metrics:
            ws_summary[f'A{row}'] = metric
            ws_summary[f'B{row}'] = value
            ws_summary[f'A{row}'].border = border
            ws_summary[f'B{row}'].border = border
            row += 1
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        
        # Interactions Detail Sheet
        ws_interactions = wb.create_sheet("Interactions")
        
        headers = ['Date', 'Client', 'Type', 'Canal', 'Statut', 'Résultat', 'Durée (min)']
        for col, header in enumerate(headers, 1):
            cell = ws_interactions.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row = 2
        for interaction in interactions:
            date_val = interaction.scheduled_date or interaction.created_at
            ws_interactions.cell(row=row, column=1, value=date_val.strftime('%d/%m/%Y') if date_val else 'N/A')
            ws_interactions.cell(row=row, column=2, value=interaction.client.get_full_name())
            ws_interactions.cell(row=row, column=3, value=interaction.get_interaction_type_display())
            ws_interactions.cell(row=row, column=4, value=interaction.get_channel_display())
            ws_interactions.cell(row=row, column=5, value=interaction.get_status_display())
            ws_interactions.cell(row=row, column=6, value=interaction.get_outcome_display() if interaction.outcome else 'N/A')
            ws_interactions.cell(row=row, column=7, value=interaction.duration_minutes or 0)
            
            for col in range(1, 8):
                ws_interactions.cell(row=row, column=col).border = border
            
            row += 1
        
        # Adjust column widths
        for col in range(1, 8):
            ws_interactions.column_dimensions[get_column_letter(col)].width = 18
        
        # Leads Detail Sheet
        ws_leads = wb.create_sheet("Leads")
        
        headers = ['Nom', 'Email', 'Téléphone', 'Source', 'Statut', 'Qualification', 'Score', 'Converti']
        for col, header in enumerate(headers, 1):
            cell = ws_leads.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        leads = Lead.objects.filter(
            assigned_agent=agent,
            created_at__gte=start_date,
            created_at__lte=end_date
        )
        
        row = 2
        for lead in leads:
            ws_leads.cell(row=row, column=1, value=lead.full_name())
            ws_leads.cell(row=row, column=2, value=lead.email)
            ws_leads.cell(row=row, column=3, value=lead.phone or 'N/A')
            ws_leads.cell(row=row, column=4, value=lead.get_source_display())
            ws_leads.cell(row=row, column=5, value=lead.get_status_display())
            ws_leads.cell(row=row, column=6, value=lead.get_qualification_display())
            ws_leads.cell(row=row, column=7, value=lead.score)
            ws_leads.cell(row=row, column=8, value='Oui' if lead.converted_to_client else 'Non')
            
            for col in range(1, 9):
                ws_leads.cell(row=row, column=col).border = border
            
            row += 1
        
        # Adjust column widths
        for col in range(1, 9):
            ws_leads.column_dimensions[get_column_letter(col)].width = 18
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generate_agency_overview_excel(agency_id, start_date=None, end_date=None):
        """
        Generate an Excel report with agency-wide performance overview.
        
        Args:
            agency_id: UUID of the agency
            start_date: Start date for the report period
            end_date: End date for the report period
            
        Returns:
            BytesIO object containing the Excel file
        """
        if not start_date:
            start_date = datetime.now() - timedelta(days=30)
        if not end_date:
            end_date = datetime.now()
        end_date = _end_of_day(end_date)
        
        try:
            agency = Agency.objects.get(id=agency_id)
        except Agency.DoesNotExist:
            return io.BytesIO()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vue d'ensemble Agence"
        
        header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=14, color="1A365D")
        
        ws['A1'] = f"Rapport d'Agence — {agency.name}"
        ws['A1'].font = title_font
        ws['A2'] = f"Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        ws.merge_cells('A1:E1')
        ws.merge_cells('A2:E2')
        
        agents = User.objects.filter(profile__agency=agency, role='agent')
        
        # Agent performance table
        ws['A4'] = "Agent"
        ws['B4'] = "Interactions"
        ws['C4'] = "Clients"
        ws['D4'] = "Leads Convertis"
        ws['E4'] = "Taux Conversion"
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}4'].font = header_font
            ws[f'{col}4'].fill = header_fill
        
        row = 5
        for agent in agents:
            interactions_count = ClientInteraction.objects.filter(
                agent=agent,
                created_at__gte=start_date,
                created_at__lte=end_date,
            ).count()
            
            clients_count = client_profiles_for_user(agent).count()
            
            leads_assigned = Lead.objects.filter(
                assigned_agent=agent,
                created_at__gte=start_date,
                created_at__lte=end_date
            ).count()
            
            leads_converted = Lead.objects.filter(
                assigned_agent=agent,
                converted_to_client=True,
                conversion_date__gte=start_date,
                conversion_date__lte=end_date
            ).count()
            
            conversion_rate = (leads_converted / leads_assigned * 100) if leads_assigned > 0 else 0
            
            ws[f'A{row}'] = agent.get_full_name()
            ws[f'B{row}'] = interactions_count
            ws[f'C{row}'] = clients_count
            ws[f'D{row}'] = leads_converted
            ws[f'E{row}'] = f"{conversion_rate:.1f}%"
            
            row += 1
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
